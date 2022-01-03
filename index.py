import pandas
import collections

from openpyxl import Workbook, load_workbook


def make_report(log_file_name, report_template_file_name, report_output_file_name):
    # Определение переменных
    browsers = {}
    browser = []
    products = ''
    products_men = ''
    products_women = ''
    products_mounth = {}
    itogo_browsers = 0

    # Открытие файла шаблона
    shablon = load_workbook(filename=report_template_file_name)
    sheet_ranges = shablon['Лист1']
    # Создание итогового файла
    wb = Workbook()
    wb["Sheet"].title = "Лист1"
    total_ws1 = wb.active
    dest_filename = report_output_file_name

    excel_data = pandas.read_excel(log_file_name, sheet_name='log', engine='openpyxl')
    excel_data_dict = excel_data.to_dict(orient='records')

    for slovar in excel_data_dict:
        name_browsers = slovar['Браузер']
        name_product = slovar['Купленные товары']
        data_visit = int(str(slovar['Дата посещения'])[5:7])

        # Вычисление браузеров по посещаемости по месяцам
        if name_browsers in browsers:
            browsers[name_browsers][data_visit-1] += 1
        else:
            browsers[name_browsers] = [0,0,0,0,0,0,0,0,0,0,0,0]
            browsers[name_browsers][data_visit-1] += 1
        #Вычисление товаров по продажам по месяцам
        for each_product in name_product.split(','):
            if each_product in products_mounth:
                products_mounth[each_product][data_visit-1] += 1
            else:
                products_mounth[each_product] = [0,0,0,0,0,0,0,0,0,0,0,0]
                products_mounth[each_product][data_visit-1] += 1
        browser.append(name_browsers)

        if slovar['Пол'] == 'м':
            products_men += f',{name_product}'
        else:
            products_women += f',{name_product}'

        #Вычисление товаров
        products += f',{name_product}'
    products_men = collections.Counter(products_men.split(','))
    products_women = collections.Counter(products_women.split(','))
    len_for_min_men = len(products_men)
    len_for_min_women = len(products_women)
    products_men_min = products_men.most_common()[len_for_min_men - 2]
    products_women_min = products_women.most_common()[len_for_min_women-2]
    products_men = products_men.most_common(1)
    products_women = products_women.most_common(1)

    #Упорядочивание товаров
    letter_product = collections.Counter(products.split(','))
    letter_product = dict(letter_product.most_common(7))
    for key, value in letter_product.items():
        for key1, value1 in products_mounth.items():
            if key == key1:
                letter_product[key] = value1

    #Упорядочивание браузеров
    letter_counter = collections.Counter(browser)
    for key, value in letter_counter.items():
        for key1, value1 in browsers.items():
            if key == key1:
                letter_counter[key] = value1

    # Запись с шаблона1
    for i in range(1, 14):
        for j in range(1, 5):
            total_ws1.cell(row=j, column=i).value = sheet_ranges.cell(row=j, column=i).value

    # Запись о браузерах в эксель
    stolb = 2
    stroka = 5

    for key, value in dict(letter_counter.most_common(7)).items():
        c2 = total_ws1.cell(row=stroka, column=1)
        c2.value = key
        for i in value:
            c1 = total_ws1.cell(row=stroka, column=stolb)
            c1.value = i
            stolb += 1
        stolb = 2
        stroka += 1
        # Запись с шаблона2
    for i in range(1, 14):
        for j in range(12, 19):
            total_ws1.cell(row=j, column=i).value = sheet_ranges.cell(row=j, column=i).value
    # Запись о товарах в эксель
    stolb = 2
    stroka = 19
    for key, value in letter_product.items():
        c2 = total_ws1.cell(row=stroka, column=1)
        c2.value = key
        for i in value:
            c1 = total_ws1.cell(row=stroka, column=stolb)
            c1.value = i
            stolb += 1
        stolb = 2
        stroka += 1

    # Запись с шаблона3
    for i in range(1, 14):
        for j in range(26, 35):
            total_ws1.cell(row=j, column=i).value = sheet_ranges.cell(row=j, column=i).value
    # Запись предпочтений
    total_ws1.cell(row=31, column=2).value = products_men[0][0]
    total_ws1.cell(row=32, column=2).value = products_women[0][0]
    total_ws1.cell(row=33, column=2).value = products_men_min[0]
    total_ws1.cell(row=34, column=2).value = products_women_min[0]
    # Запись ИТОГО
    for i in range(2, 14):
        for j in range(5, 12):
            itogo_browsers += total_ws1.cell(row=j, column=i).value
        total_ws1.cell(row=12, column=i).value = itogo_browsers
        itogo_browsers = 0
    for i in range(2, 14):
        for j in range(19, 26):
            itogo_browsers += total_ws1.cell(row=j, column=i).value
        total_ws1.cell(row=26, column=i).value = itogo_browsers
        itogo_browsers = 0

    # Сохранение файла
    wb.save(dest_filename)
    pass

